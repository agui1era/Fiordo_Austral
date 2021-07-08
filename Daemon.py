
import os
import datetime
import math
import psycopg2
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from ftplib import FTP


#S/E 300 KVA              jt9DdWAvNK            d37982b0-c220-11eb-a61d-e9bafc595a10
#Desodorizado Omega 3     HDjYvFtfZ7            ed91f600-c220-11eb-a61d-e9bafc595a10
#WINT. VEGETALES          WEbn35kQRk            01d83cf0-c221-11eb-a61d-e9bafc595a10
#S/E 500                  OuvVsF0vek            c2375400-c220-11eb-a61d-e9bafc595a10
#REF. VEGETALES           eo085EanGe            4b50ade0-c221-11eb-a61d-e9bafc595a10
#CALDERA                  hmqPNgajnY            5da0a310-c221-11eb-a61d-e9bafc595a10
#DESO. VEGETALES          gQVSvMsyJt            8adc3a10-c221-11eb-a61d-e9bafc595a10
#REF. OMEGA 3             5Msyx2m8Oh            0aa404d0-c222-11eb-a61d-e9bafc595a10
#S/E 1000                 eBSXNYdrl9            b3a243f0-c220-11eb-a61d-e9bafc595a10
#PRENSAS                  SBslPQTJAa            e70ed3b0-c221-11eb-a61d-e9bafc595a10 
#Extracción               UBpnpsCWB8            fcc01c50-c221-11eb-a61d-e9bafc595a10         
#TRIOMAX                  UpekFmGS3w            1d869d10-c222-11eb-a61d-e9bafc595a10
#CHI. TRIOMAX             K2HcV0IHY6            9971bf50-c221-11eb-a61d-e9bafc595a10
#CHI. VEGETALES           BoJd2ibLrC            ffe25760-d39a-11eb-a594-afa0469684b4


def getDB(sql_query):
    try:       
        connection = psycopg2.connect(user = "postgres",
                                        password = "imagina12",
                                        host = "iot.igromi.com",
                                        port = "5432",
                                        database = "thingsboard")
        cursor = connection.cursor()
        postgreSQL_select_Query = sql_query

        cursor.execute(postgreSQL_select_Query)
        bd_records = cursor.fetchall()
        for row in bd_records:
            i=1

    except (Exception, psycopg2.Error) as error:
        print("Error fetching data from PostgreSQL table", error)
    finally:
        # closing database connection
        if (connection):
            cursor.close()
            connection.close()
    try:
       
       out_query=row[0]

    except (Exception, psycopg2.Error) as error:
       out_query=0

    return out_query

def date_to_milis(date_string):

    #convert date to timestamp
    obj_date = datetime.datetime.strptime(date_string,"%d/%m/%Y %H:%M:%S")

    return str(math.trunc(obj_date.timestamp() * 1000))

end_date = datetime.datetime.now()
str_end_date=end_date.strftime("30/%m/%Y 23:59:59")
begin_date = end_date  - relativedelta(months=1)
str_begin_date=begin_date.strftime("30/%m/%Y 23:59:59")
print(str_begin_date)

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='d37982b0-c220-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_SE300=getDB(sql_str_det)
print('Energia mes anterior SE300 '+str(ENERGIA_ANT_SE300))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='ed91f600-c220-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_DESO_OMEGA3=getDB(sql_str_det)
print('Energia mes anterior DESO_OMEGA3 '+str(ENERGIA_ANT_DESO_OMEGA3))  

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='01d83cf0-c221-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_WINT_VEGETALES=getDB(sql_str_det)
print('Energia mes anterior WINT_VEGETALES '+str(ENERGIA_ANT_WINT_VEGETALES)) 

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='c2375400-c220-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_SE500=getDB(sql_str_det)
print('Energia mes anterior  SE500 '+str(ENERGIA_ANT_SE500))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='4b50ade0-c221-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_REF_VEGETALES=getDB(sql_str_det)
print('Energia mes anterior  REF_VEGETALES '+str(ENERGIA_ANT_REF_VEGETALES))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='5da0a310-c221-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_CALDERA=getDB(sql_str_det)
print('Energia mes anterior  CALDERA '+str(ENERGIA_ANT_CALDERA)) 

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='8adc3a10-c221-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_DESO_VEGETALES=getDB(sql_str_det)
print('Energia mes anterior  DESO_VEGETALES '+str(ENERGIA_ANT_DESO_VEGETALES))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='0aa404d0-c222-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_REF_OMEGA3=getDB(sql_str_det)
print('Energia mes anterior  REF_OMEGA3 '+str(ENERGIA_ANT_REF_OMEGA3))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='b3a243f0-c220-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_SE1000=getDB(sql_str_det)
print('Energia mes anterior  SE1000 '+str(ENERGIA_ANT_SE1000))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='e70ed3b0-c221-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_PRENSAS=getDB(sql_str_det)
print('Energia mes anterior  PRENSAS '+str(ENERGIA_ANT_PRENSAS))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='9971bf50-c221-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_CHI_TRIOMAX=getDB(sql_str_det)
print('Energia mes anterior  CHI_TRIOMAX '+str(ENERGIA_ANT_CHI_TRIOMAX))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='fcc01c50-c221-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_EXTRACCION=getDB(sql_str_det)
print('Energia mes anterior  PRENSAS '+str(ENERGIA_ANT_EXTRACCION))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='1d869d10-c222-11eb-a61d-e9bafc595a10' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_TRIOMAX=getDB(sql_str_det)
print('Energia mes anterior  TRIOMAX '+str(ENERGIA_ANT_TRIOMAX))

sql_str_det="SELECT long_v FROM ts_kv WHERE ts <= "+ date_to_milis(str_begin_date)+ " AND key=74 " + "AND  entity_id='ffe25760-d39a-11eb-a594-afa0469684b' ORDER BY ts DESC LIMIT 1"
ENERGIA_ANT_CHI_VEGETALES=getDB(sql_str_det)
print('Energia mes anterior  CHI_VEGETALES '+str(ENERGIA_ANT_CHI_VEGETALES))

token_dispositivos = ['jt9DdWAvNK',         'HDjYvFtfZ7',                  'WEbn35kQRk',                      'OuvVsF0vek' ,        'eo085EanGe',                   'hmqPNgajnY',           'gQVSvMsyJt',                   '5Msyx2m8Oh',                   'eBSXNYdrl9',           'SBslPQTJAa',            'UBpnpsCWB8',          'UpekFmGS3w',           'K2HcV0IHY6',          'BoJd2ibLrC']
energia_anterior =   [ENERGIA_ANT_SE300,    ENERGIA_ANT_DESO_OMEGA3,        ENERGIA_ANT_WINT_VEGETALES,       ENERGIA_ANT_SE500,    ENERGIA_ANT_REF_VEGETALES,      ENERGIA_ANT_CALDERA,    ENERGIA_ANT_DESO_VEGETALES,     ENERGIA_ANT_REF_OMEGA3,         ENERGIA_ANT_SE1000,    ENERGIA_ANT_PRENSAS, ENERGIA_ANT_EXTRACCION,   ENERGIA_ANT_TRIOMAX,   ENERGIA_ANT_CHI_TRIOMAX,   ENERGIA_ANT_CHI_VEGETALES]



file='ultimo.xlsx'
ftp = FTP('igromi.com')
ftp.login('igromi','diccionarioAvanza12')

ftp.retrbinary("RETR /fiordo/ultimo(01).xlsx" ,open( file, 'wb').write)

ftp.quit()

token_diff='jzEVcu4ocZ'
token_totales='X6R56MUqQl'

wb = load_workbook(file)
wb.iso_dates = True
sheet = wb['||||Hoja111']

Columna=1

MAX_FILAS=sheet.max_row
print('Filas totales: '+str(MAX_FILAS))

time_now = datetime.datetime.now()
begin_time = time_now  - datetime.timedelta(hours=12)

for Fila in range(MAX_FILAS+4):
    Columna=1
    Fila=Fila+1
    FECHA=sheet.cell(row=Fila+4, column=1).value
    
    print('------------------------------------------------------')
    print('HORA PLANILLA : '+str(FECHA))
    print(' > HORA ACTUAL MENOS DELTA HORAS: '+str(begin_time))
    print('------------------------------------------------------')
    
    try:     
        if (FECHA > begin_time) :  

            TIMESTAMP=datetime.datetime.strptime(str(FECHA), '%Y-%m-%d %H:%M:%S').timestamp()*1000
            SE300=sheet.cell(row=Fila+4, column=2).value
            DESO_OMEGA3=sheet.cell(row=Fila+4, column=9).value
            WINT_VEGETALES=sheet.cell(row=Fila+4, column=16).value
            SE500=sheet.cell(row=Fila+4, column=23).value
            REF_VEGETALES=sheet.cell(row=Fila+4, column=30).value
            CALDERA=sheet.cell(row=Fila+4, column=37).value
            DESO_VEGETALES=sheet.cell(row=Fila+4, column=44).value
            REF_OMEGA3=sheet.cell(row=Fila+4, column=51).value
            SE1000=sheet.cell(row=Fila+4, column=58).value
            PRENSAS=sheet.cell(row=Fila+4, column=65).value
            EXTRACCION=sheet.cell(row=Fila+4, column=72).value
            TRIOMAX=sheet.cell(row=Fila+4, column=79).value
            CHI_TRIOMAX=sheet.cell(row=Fila+4, column=86).value
            CHI_VEGETALES=sheet.cell(row=Fila+4, column=93).value

            DIFF_SE300 = SE300-DESO_OMEGA3-WINT_VEGETALES
            DIFF_SE500 = SE500-REF_VEGETALES-CALDERA-DESO_VEGETALES
            DIFF_SE1000= SE1000-PRENSAS-EXTRACCION-TRIOMAX-CHI_TRIOMAX-REF_OMEGA3-CHI_VEGETALES

            TOTAL_TRIOMAX=TRIOMAX+CHI_TRIOMAX
            TOTAL_OMEGA=DESO_OMEGA3+REF_OMEGA3
            TOTAL_VEGETALES=WINT_VEGETALES+REF_VEGETALES+DESO_VEGETALES+PRENSAS+EXTRACCION+CHI_VEGETALES

            TOTAL_SE=SE300+SE500+SE1000

            PORCENT_TRIOMAX=(TOTAL_TRIOMAX/TOTAL_SE)*100
            PORCENT_OMEGA=(TOTAL_OMEGA/TOTAL_SE)*100
            PORCENT_VEGETALES=(TOTAL_VEGETALES/TOTAL_SE)*100
            PORCENT_OTROS300=(DIFF_SE300/TOTAL_SE)*100
            PORCENT_OTROS500=(DIFF_SE500/TOTAL_SE)*100
            PORCENT_OTROS1000=(DIFF_SE1000/TOTAL_SE)*100
            PORCENT_CALDERA=(CALDERA/TOTAL_SE)*100
            PORCENT_SE300=(SE300/TOTAL_SE)*100
            PORCENT_SE500=(SE500/TOTAL_SE)*100
            PORCENT_SE1000=(SE1000/TOTAL_SE)*100

            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"DIFF_SE300\":'+str(DIFF_SE300)+'}}" iot.igromi.com:8080/api/v1/'+token_diff+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"DIFF_SE500\":'+str(DIFF_SE500)+'}}" iot.igromi.com:8080/api/v1/'+token_diff+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"DIFF_SE1000\":'+str(DIFF_SE1000)+'}}" iot.igromi.com:8080/api/v1/'+token_diff+'/telemetry --header "Content-Type:application/json"')

            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_TRIOMAX\":'+str(PORCENT_TRIOMAX)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_OMEGA\":'+str(PORCENT_OMEGA)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_VEGETALES\":'+str(PORCENT_VEGETALES)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_OTROS300\":'+str(PORCENT_OTROS300)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_OTROS500\":'+str(PORCENT_OTROS500)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_OTROS1000\":'+str(PORCENT_OTROS1000)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_CALDERA\":'+str(PORCENT_CALDERA)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_SE300":'+str(PORCENT_SE300)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_SE500\":'+str(PORCENT_SE500)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"PORCENT_SE1000\":'+str(PORCENT_SE1000)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_TRIOMAX\":'+str(ENERGIA_ANT_TRIOMAX)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_CHI_TRIOMAX":'+str(ENERGIA_ANT_CHI_TRIOMAX)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_REF_OMEGA3\":'+str(ENERGIA_ANT_REF_OMEGA3)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')           
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_DESO_OMEGA3\":'+str(ENERGIA_ANT_DESO_OMEGA3)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"') 
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_SE300":'+str(ENERGIA_ANT_SE300)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_SE500\":'+str(ENERGIA_ANT_SE500)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_SE1000\":'+str(ENERGIA_ANT_SE1000)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_CHI_VEGETALES\":'+str(ENERGIA_ANT_CHI_VEGETALES)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_WINT_VEGETALES\":'+str( ENERGIA_ANT_WINT_VEGETALES)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_REF_VEGETALES\":'+str(ENERGIA_ANT_REF_VEGETALES)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_EXTRACCION\":'+str( ENERGIA_ANT_EXTRACCION)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_PRENSA\":'+str(ENERGIA_ANT_CALDERA)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"')
            os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"ANT_CALDERA\":'+str(ENERGIA_ANT_PRENSAS)+'}}" iot.igromi.com:8080/api/v1/'+token_totales+'/telemetry --header "Content-Type:application/json"') 


             
            indice=0

            for token in token_dispositivos:
                
                POT=sheet.cell(row=Fila+4, column=1+Columna).value-energia_anterior[indice]/10
                IA=sheet.cell(row=Fila+4, column=2+Columna).value
                IB=sheet.cell(row=Fila+4, column=3+Columna).value
                IC=sheet.cell(row=Fila+4, column=4+Columna).value
                VAB=sheet.cell(row=Fila+4, column=5+Columna).value
                VAC=sheet.cell(row=Fila+4, column=6+Columna).value
                VBC=sheet.cell(row=Fila+4, column=7+Columna).value

                indice=indice+1
                  
                print('------------------------------------------------------')
                print('FILA:' +str(Fila+4))
                print('FECHA: '+str(FECHA))
                print('POT: '+str(POT))
                print('IA: '+str(IA))
                print('IB: '+str(IB))
                print('IC: '+str(IC))
                print('VAB: '+str(VAB))
                print('VAC: '+str(VAC))
                print('VBC: '+str(VBC))
                print('------------------------------------------------------')

                Columna=Columna+7

                os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"POT\":'+str(POT)+'}}" iot.igromi.com:8080/api/v1/'+token+'/telemetry --header "Content-Type:application/json"')
                os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"IA\":'+str(IA)+'}}" iot.igromi.com:8080/api/v1/'+token+'/telemetry --header "Content-Type:application/json"')
                os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"IB\":'+str(IB)+'}}" iot.igromi.com:8080/api/v1/'+token+'/telemetry --header "Content-Type:application/json"')
                os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"IC\":'+str(IC)+'}}" iot.igromi.com:8080/api/v1/'+token+'/telemetry --header "Content-Type:application/json"')
                os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"VAB\":'+str(VAB)+'}}" iot.igromi.com:8080/api/v1/'+token+'/telemetry --header "Content-Type:application/json"')
                os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"VAC\":'+str(VAC)+'}}" iot.igromi.com:8080/api/v1/'+token+'/telemetry --header "Content-Type:application/json"')
                os.system('curl -v -X POST -d "{\"ts\":'+str(TIMESTAMP)+',\"values\":{\"VBC\":'+str(VBC)+'}}" iot.igromi.com:8080/api/v1/'+token+'/telemetry --header "Content-Type:application/json"')

    except:
        print('error')